"""
生态导出模块
============
支持导出为: Obsidian Markdown / EPUB / Excel / GEXF

用法:
    exporter = NovelExporter("shaosong")
    exporter.export_obsidian("output/obsidian/")
    exporter.export_epub("output/shaosong.epub")
    exporter.export_excel("output/shaosong.xlsx")
    exporter.export_markdown_report("output/report.md")
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 项目路径
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
WIKI_DIR = DATA_DIR / "wiki"


# ── 工具函数 ────────────────────────────────────────────────────────────

def _safe_filename(s: str) -> str:
    """将字符串转为安全的文件名"""
    return re.sub(r'[\\/:*?"<>|]', "_", s)


# ── 数据加载 ────────────────────────────────────────────────────────────

def load_novel_data(novel_key: str) -> dict | None:
    """加载一部小说的所有数据"""
    wiki_path = WIKI_DIR / f"{novel_key}_hierarchical.json"
    graph_path = WIKI_DIR / f"{novel_key}_graph.json"

    data = {"key": novel_key, "wiki": {}, "graph": {"nodes": [], "edges": []}}

    if wiki_path.exists():
        with open(wiki_path, "r", encoding="utf-8") as f:
            data["wiki"] = json.load(f)

    if graph_path.exists():
        with open(graph_path, "r", encoding="utf-8") as f:
            data["graph"] = json.load(f)

    return data


# ── 主导出器 ────────────────────────────────────────────────────────────

class NovelExporter:
    """小说数据导出器"""

    def __init__(self, novel_key: str):
        self.key = novel_key
        self.data = load_novel_data(novel_key)
        self.wiki = self.data.get("wiki", {})
        self.graph = self.data.get("graph", {})
        self.chapters = self.wiki.get("chapters", []) if isinstance(self.wiki, dict) else []
        self.volumes = self.wiki.get("volumes", []) if isinstance(self.wiki, dict) else []
        self.book = self.wiki.get("book", {}) if isinstance(self.wiki, dict) else {}

        cn_map = {"shaosong": "绍宋", "斗破苍穹": "斗破苍穹", "神印王座": "神印王座"}
        self.display_name = cn_map.get(novel_key, novel_key)

    # ═════════════════════════════════════════════════════════════════
    #  Obsidian 知识库导出
    # ═════════════════════════════════════════════════════════════════

    def export_obsidian(self, output_dir: str) -> dict:
        """
        导出为 Obsidian 知识库（双向链接 Markdown）。

        输出结构:
            output_dir/
            ├── 全书总览.md
            ├── 人物/
            │   ├── 赵玖.md
            │   ├── 岳飞.md
            │   └── ...
            ├── 卷/
            │   ├── 第1-50章.md
            │   └── ...
            └── 章节/
                ├── 第一章 明道宫.md
                └── ...
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # 子目录
        char_dir = output_path / "人物"
        vol_dir = output_path / "卷"
        ch_dir = output_path / "章节"
        for d in [char_dir, vol_dir, ch_dir]:
            d.mkdir(exist_ok=True)

        stats = {"characters": 0, "volumes": 0, "chapters": 0}

        # 1. 全书总览
        if self.book:
            book_content = self._obsidian_book_page()
            (output_path / "全书总览.md").write_text(book_content, encoding="utf-8")

        # 2. 人物页面
        nodes = self.graph.get("nodes", [])
        edges = self.graph.get("edges", [])
        # 构建邻接表
        adj: dict[str, list[dict]] = {}
        for e in edges:
            s, t = e["source"], e["target"]
            adj.setdefault(s, []).append(e)
            adj.setdefault(t, []).append(e)

        for n in nodes:
            name = n["name"]
            rels = adj.get(name, [])
            # 找出相关章节
            chap_titles = set()
            for r in rels:
                for ch in r.get("chapters", []):
                    chap_titles.add(ch)

            content = self._obsidian_character_page(name, n, rels, list(chap_titles))
            fn = _safe_filename(name) + ".md"
            (char_dir / fn).write_text(content, encoding="utf-8")
            stats["characters"] += 1

        # 3. 卷页面
        for vol in self.volumes:
            content = self._obsidian_volume_page(vol)
            fn = _safe_filename(vol.get("title", "")) + ".md"
            (vol_dir / fn).write_text(content, encoding="utf-8")
            stats["volumes"] += 1

        # 4. 章节页面
        for ch in self.chapters:
            content = self._obsidian_chapter_page(ch)
            title = ch.get("chapter_title", f"第{ch.get('chapter_index', 0)}章")
            fn = _safe_filename(title) + ".md"
            (ch_dir / fn).write_text(content, encoding="utf-8")
            stats["chapters"] += 1

        logger.info(
            f"[Obsidian] 导出完成: {output_dir} "
            f"({stats['characters']} 人物, {stats['volumes']} 卷, {stats['chapters']} 章)"
        )
        return stats

    def _obsidian_book_page(self) -> str:
        """生成全书总览页面"""
        lines = [
            f"# {self.display_name}",
            f"",
            f"{self.book.get('summary', '')}",
            f"",
            f"## 主要人物",
        ]
        for mc in self.book.get("main_characters", []):
            lines.append(f"- [[人物/{_safe_filename(mc)}|{mc}]]")
        lines.extend(["", "## 主题", ""])
        for theme in self.book.get("themes", []):
            lines.append(f"- {theme}")
        lines.extend(["", "## 卷", ""])
        for vol in self.volumes:
            title = vol.get("title", "")
            lines.append(f"- [[卷/{_safe_filename(title)}|{title}]]")
        return "\n".join(lines)

    def _obsidian_character_page(self, name: str, node: dict, rels: list[dict], chapters: list[str]) -> str:
        """生成人物页面"""
        lines = [
            f"# {name}",
            f"",
            f"**角色**: {node.get('role', '未知')}  ",
            f"**出场次数**: {node.get('mention_count', 0)}  ",
            f"**相关章节**: {node.get('chapter_count', 0)}",
            f"",
            f"## 关系网络",
            f"",
        ]
        for r in rels:
            partner = r["target"] if r["source"] == name else r["source"]
            lines.append(f"- [[人物/{_safe_filename(partner)}|{partner}]] — {r.get('relation', '')}")
        lines.extend(["", "## 出场章节", ""])
        for ch_title in chapters[:20]:
            lines.append(f"- [[章节/{_safe_filename(ch_title)}|{ch_title}]]")
        if len(chapters) > 20:
            lines.append(f"- ...及另外 {len(chapters) - 20} 章")
        return "\n".join(lines)

    def _obsidian_volume_page(self, vol: dict) -> str:
        """生成卷页面"""
        lines = [
            f"# {vol.get('title', '')}",
            f"",
            f"**章节范围**: 第{vol['chapter_range'][0]}-{vol['chapter_range'][1]}章  ",
            f"",
            f"{vol.get('summary', '')}",
            f"",
            f"## 主要人物",
        ]
        for mc in vol.get("main_characters", []):
            # 去掉括号备注
            clean_name = mc.split("（")[0].split("(")[0]
            lines.append(f"- [[人物/{_safe_filename(clean_name)}|{mc}]]")
        return "\n".join(lines)

    def _obsidian_chapter_page(self, ch: dict) -> str:
        """生成章节页面"""
        title = ch.get("chapter_title", "")
        idx = ch.get("chapter_index", 0)
        lines = [
            f"# {title}",
            f"",
            f"**索引**: {idx}",
            f"",
            f"## 摘要",
            f"{ch.get('summary', '')}",
            f"",
        ]
        chars = ch.get("characters", [])
        if chars:
            lines.extend(["## 出场人物", ""])
            for c in chars:
                name = c["name"]
                desc = c.get("description", "")
                lines.append(f"- [[人物/{_safe_filename(name)}|{name}]]: {desc}")
            lines.append("")
        events = ch.get("events", [])
        if events:
            lines.extend(["## 关键事件", ""])
            for ev in events:
                lines.append(f"- {ev}")
            lines.append("")
        return "\n".join(lines)

    # ═════════════════════════════════════════════════════════════════
    #  EPUB 导出
    # ═════════════════════════════════════════════════════════════════

    def export_epub(self, output_path: str) -> bool:
        """
        导出为 EPUB 电子书。

        需要安装: pip install ebooklib
        如果未安装，会回退为生成 HTML 文件。
        """
        try:
            from ebooklib import epub
        except ImportError:
            logger.warning("ebooklib 未安装，回退为 HTML 导出")
            html_path = output_path.replace(".epub", ".html")
            return self.export_html(html_path)

        book = epub.EpubBook()
        book.set_identifier(f"novel-graphrag-{self.key}")
        book.set_title(self.display_name)
        book.set_language("zh-CN")

        # CSS
        style = """
        body { font-family: "Source Han Serif", "Noto Serif CJK SC", serif; line-height: 1.8; padding: 1em; }
        h1 { color: #8B0000; border-bottom: 2px solid #8B0000; }
        h2 { color: #444; }
        .summary { font-style: italic; color: #666; }
        """
        css = epub.EpubItem(
            uid="style",
            file_name="style/nav.css",
            media_type="text/css",
            content=style,
        )
        book.add_item(css)

        chapters_epub = []
        # 全书摘要
        if self.book and self.book.get("summary"):
            c = epub.EpubHtml(
                title="全书总览",
                file_name="book.xhtml",
                lang="zh-CN",
            )
            c.content = f"<h1>{self.display_name}</h1><p class='summary'>{self.book['summary']}</p>"
            book.add_item(c)
            chapters_epub.append(c)

        # 逐章
        for i, ch in enumerate(self.chapters):
            title = ch.get("chapter_title", f"Chapter {i}")
            summary = ch.get("summary", "")
            chars = ch.get("characters", [])

            html = [f"<h2>{title}</h2>", f"<p class='summary'>{summary}</p>"]
            if chars:
                html.append("<h3>人物</h3><ul>")
                for c in chars:
                    html.append(f"<li><b>{c['name']}</b> ({c.get('role', '')}): {c.get('description', '')}</li>")
                html.append("</ul>")

            c = epub.EpubHtml(title=title, file_name=f"ch_{i}.xhtml", lang="zh-CN")
            c.content = "\n".join(html)
            book.add_item(c)
            chapters_epub.append(c)

        # 目录
        book.toc = chapters_epub
        book.add_item(epub.EpubNcx())
        book.add_item(epub.EpubNav())
        book.spine = chapters_epub

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        epub.write_epub(output_path, book)
        logger.info(f"[EPUB] 导出完成: {output_path} ({len(chapters_epub)} 章)")
        return True

    def export_html(self, output_path: str) -> bool:
        """导出为单个 HTML 文件（EPUB 的降级方案）"""
        parts = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'>",
            f"<title>{self.display_name}</title>",
            "<style>body{font-family:serif;line-height:1.8;max-width:800px;margin:auto;padding:2em}"
            "h1{color:#8B0000;border-bottom:2px solid #8B0000}"
            ".summary{color:#666;font-style:italic}</style></head><body>",
            f"<h1>{self.display_name}</h1>",
        ]
        if self.book and self.book.get("summary"):
            parts.append(f"<p class='summary'>{self.book['summary']}</p><hr>")

        for ch in self.chapters:
            title = ch.get("chapter_title", "")
            summary = ch.get("summary", "")
            parts.append(f"<h2>{title}</h2><p class='summary'>{summary}</p>")
            for c in ch.get("characters", []):
                parts.append(f"<p><b>{c['name']}</b>: {c.get('description', '')}</p>")

        parts.append("</body></html>")
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(parts))
        logger.info(f"[HTML] 导出完成: {output_path}")
        return True

    # ═════════════════════════════════════════════════════════════════
    #  Excel 人物清单导出
    # ═════════════════════════════════════════════════════════════════

    def export_excel(self, output_path: str) -> bool:
        """
        导出人物清单 + 关系清单为 Excel。

        需要安装: pip install openpyxl
        如果未安装，回退为 CSV。
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            logger.warning("openpyxl 未安装，回退为 CSV")
            return self.export_csv(output_path.replace(".xlsx", ".csv"))

        wb = openpyxl.Workbook()

        # Sheet 1: 人物
        ws1 = wb.active
        ws1.title = "人物"
        ws1.append(["姓名", "角色", "出场次数", "相关章节数"])
        for c in ws1.iter_cols(min_col=1, max_col=4):
            c[0].font = Font(bold=True)

        for n in self.graph.get("nodes", []):
            ws1.append([
                n.get("name", ""),
                n.get("role", ""),
                n.get("mention_count", 0),
                n.get("chapter_count", 0),
            ])

        # Sheet 2: 关系
        ws2 = wb.create_sheet("关系")
        ws2.append(["人物 A", "关系描述", "人物 B", "权重", "出现章节数"])
        for c in ws2.iter_cols(min_col=1, max_col=5):
            c[0].font = Font(bold=True)

        for e in self.graph.get("edges", []):
            ws2.append([
                e.get("source", ""),
                e.get("relation", ""),
                e.get("target", ""),
                e.get("weight", 1),
                len(e.get("chapters", [])),
            ])

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info(f"[Excel] 导出完成: {output_path}")
        return True

    def export_csv(self, output_path: str) -> bool:
        """回退为 CSV 导出"""
        import csv
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        base = output_path.replace(".csv", "")

        # 人物 CSV
        with open(f"{base}_characters.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["姓名", "角色", "出场次数", "章节数"])
            for n in self.graph.get("nodes", []):
                w.writerow([n.get("name"), n.get("role"), n.get("mention_count"), n.get("chapter_count")])

        # 关系 CSV
        with open(f"{base}_relations.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["人物A", "关系", "人物B", "权重"])
            for e in self.graph.get("edges", []):
                w.writerow([e.get("source"), e.get("relation"), e.get("target"), e.get("weight")])

        logger.info(f"[CSV] 导出完成: {base}_*.csv")
        return True

    # ═════════════════════════════════════════════════════════════════
    #  Markdown 分析报告导出
    # ═════════════════════════════════════════════════════════════════

    def export_markdown_report(self, output_path: str) -> bool:
        """导出全书分析报告为 Markdown"""
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        nodes = self.graph.get("nodes", [])
        edges = self.graph.get("edges", [])

        # 统计
        roles = {}
        for n in nodes:
            role = n.get("role", "未知")
            roles[role] = roles.get(role, 0) + 1

        top_chars = sorted(nodes, key=lambda n: -n.get("mention_count", 0))[:10]
        top_edges = sorted(edges, key=lambda e: -e.get("weight", 0))[:10]

        lines = [
            "# " + "《" + self.display_name + "》分析报告",
            f"",
            f"> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"> 数据来源: Novel-GraphRAG",
            f"",
            f"---",
            f"",
            f"## 1. 全书概览",
            f"",
            f"{self.book.get('summary', '暂无摘要')}",
            f"",
            f"## 2. 人物图谱统计",
            f"",
            f"| 指标 | 数值 |",
            f"|------|------|",
            f"| 总人物数 | {len(nodes)} |",
            f"| 总关系数 | {len(edges)} |",
            f"| 总章节数 | {len(self.chapters)} |",
            f"| 总卷数 | {len(self.volumes)} |",
            f"",
            f"### 角色分布",
            f"",
            f"| 角色类型 | 数量 |",
            f"|----------|------|",
        ]
        for role, count in sorted(roles.items(), key=lambda x: -x[1]):
            lines.append(f"| {role} | {count} |")

        lines.extend([
            "",
            "## 3. 主要人物 Top 10",
            "",
            "| 排名 | 姓名 | 角色 | 出场次数 |",
            "|------|------|------|----------|",
        ])
        for i, n in enumerate(top_chars, 1):
            lines.append(f"| {i} | {n['name']} | {n.get('role', '')} | {n.get('mention_count', 0)} |")

        lines.extend([
            "",
            "## 4. 核心关系 Top 10",
            "",
            "| 人物 A | 关系 | 人物 B | 权重 |",
            "|--------|------|--------|------|",
        ])
        for e in top_edges:
            lines.append(f"| {e['source']} | {e['relation'][:30]} | {e['target']} | {e.get('weight', 0)} |")

        # 卷摘要
        if self.volumes:
            lines.extend(["", "## 5. 卷摘要", ""])
            for vol in self.volumes:
                cr = vol.get("chapter_range", [0, 0])
                lines.extend([
                    f"### {vol.get('title')} （第{cr[0]}-{cr[1]}章）",
                    f"{vol.get('summary', '')}",
                    "",
                ])

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        logger.info(f"[Markdown] 报告导出完成: {output_path}")
        return True


# ── 命令行 ──────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Novel-GraphRAG 导出工具")
    parser.add_argument("--novel", default="shaosong", help="小说 key")
    parser.add_argument("--output", default="output", help="输出目录或文件")
    parser.add_argument("--format", choices=["obsidian", "epub", "html", "excel", "csv", "report"],
                       default="report", help="导出格式")
    args = parser.parse_args()

    exporter = NovelExporter(args.novel)
    output = args.output

    if args.format == "obsidian":
        stats = exporter.export_obsidian(output)
        print(f"Obsidian 导出完成: {stats}")
    elif args.format == "epub":
        exporter.export_epub(f"{output}/{args.novel}.epub")
    elif args.format == "html":
        exporter.export_html(f"{output}/{args.novel}.html")
    elif args.format == "excel":
        exporter.export_excel(f"{output}/{args.novel}.xlsx")
    elif args.format == "csv":
        exporter.export_csv(f"{output}/{args.novel}.csv")
    else:
        exporter.export_markdown_report(f"{output}/{args.novel}_report.md")


if __name__ == "__main__":
    main()
